import json
import os
import sys
import getpass
from datetime import datetime

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor, QBrush
from PyQt5.QtWidgets import (
    QApplication, QFileDialog, QComboBox, QFrame, QHBoxLayout, QHeaderView,
    QLabel, QMainWindow, QMessageBox, QPushButton, QStyledItemDelegate,
    QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side


# =========================
# 환경 설정
# =========================
SHARED_FOLDER = r"C:\Shared\financial_request_data"
DEFAULT_WRITE_ROWS = 12

COLUMNS = [
    "순번", "문서번호", "요청기관", "실명번호", "계좌번호",
    "거래내역", "인적사항", "SMS정보", "거래신청서", "사고신고내역", "담당자명"
]

TEXT_COLUMNS = [1, 2, 3, 4, 10]
CHECK_COLUMNS = [5, 6, 7, 8, 9]

CHECK_OFF = "[ ]"
CHECK_ON = "[✔]"

CHECK_COLUMN_NAME_MAP = {
    5: "거래내역",
    6: "인적사항",
    7: "SMS정보",
    8: "거래신청서",
    9: "사고신고내역",
}

HEADER_INDEX_MAP = {
    "문서번호": 1,
    "요청기관": 2,
    "실명번호": 3,
    "계좌번호": 4,
    "거래내역": 5,
    "인적사항": 6,
    "SMS정보": 7,
    "거래신청서": 8,
    "사고신고내역": 9,
    "담당자명": 10,
}

THEMES = {
    "민트 아쿠아": {
        "bg_main": "#EAFBF6",
        "bg_panel": "#F5FFFC",
        "bg_tab": "#DFFBF1",
        "bg_tab_selected": "#B8F3E0",
        "bg_button": "#7FFFD4",
        "bg_button_hover": "#66EFC4",
        "bg_button_pressed": "#4FD9B1",
        "bg_header": "#D8F9EE",
        "text_main": "#1F2937",
        "text_title": "#145E55",
        "table_select": "#D6FFF4",
        "row_no_bg": "#F1FFFA",
        "child_dim_bg": "#F7FFFC",
        "child_dim_fg": "#7C8F8A",
    },
    "블러시 핑크": {
        "bg_main": "#FFF3F7",
        "bg_panel": "#FFF9FB",
        "bg_tab": "#FFE6EE",
        "bg_tab_selected": "#FDD7E4",
        "bg_button": "#FDD7E4",
        "bg_button_hover": "#F9C1D4",
        "bg_button_pressed": "#F4A8C0",
        "bg_header": "#FFE5EE",
        "text_main": "#2F2A32",
        "text_title": "#8C3A5B",
        "table_select": "#FFEAF1",
        "row_no_bg": "#FFF8FB",
        "child_dim_bg": "#FFF9FB",
        "child_dim_fg": "#99828C",
    },
    "로즈 핑크": {
        "bg_main": "#FFF1F6",
        "bg_panel": "#FFF8FB",
        "bg_tab": "#FFD7E5",
        "bg_tab_selected": "#F778A1",
        "bg_button": "#F778A1",
        "bg_button_hover": "#EC628F",
        "bg_button_pressed": "#DB4E7D",
        "bg_header": "#FFDCE8",
        "text_main": "#2F2330",
        "text_title": "#8A2F56",
        "table_select": "#FFE3ED",
        "row_no_bg": "#FFF8FB",
        "child_dim_bg": "#FFF9FB",
        "child_dim_fg": "#9A7E8A",
    },
    "아이스 블루": {
        "bg_main": "#F0FAFF",
        "bg_panel": "#F8FDFF",
        "bg_tab": "#DDF4FF",
        "bg_tab_selected": "#BDEDFF",
        "bg_button": "#BDEDFF",
        "bg_button_hover": "#A4E1FB",
        "bg_button_pressed": "#88D2F2",
        "bg_header": "#DFF3FF",
        "text_main": "#1F2937",
        "text_title": "#2E6685",
        "table_select": "#E6F8FF",
        "row_no_bg": "#F6FCFF",
        "child_dim_bg": "#FAFEFF",
        "child_dim_fg": "#7B8F99",
    },
    "라벤더": {
        "bg_main": "#F6F3FF",
        "bg_panel": "#FCFBFF",
        "bg_tab": "#EAE4FF",
        "bg_tab_selected": "#CDBDFF",
        "bg_button": "#B9A7FF",
        "bg_button_hover": "#A892FA",
        "bg_button_pressed": "#9277EA",
        "bg_header": "#EEE8FF",
        "text_main": "#26263A",
        "text_title": "#5A4CA0",
        "table_select": "#EFE9FF",
        "row_no_bg": "#FBFAFF",
        "child_dim_bg": "#FCFBFF",
        "child_dim_fg": "#857D9B",
    },
    "네이비 스카이": {
        "bg_main": "#EEF5FF",
        "bg_panel": "#F8FBFF",
        "bg_tab": "#DCE9FF",
        "bg_tab_selected": "#AFCBFF",
        "bg_button": "#7DA8FF",
        "bg_button_hover": "#6998F5",
        "bg_button_pressed": "#5584E0",
        "bg_header": "#DFEAFF",
        "text_main": "#1F2937",
        "text_title": "#2D4F8F",
        "table_select": "#E5EEFF",
        "row_no_bg": "#F8FBFF",
        "child_dim_bg": "#FBFDFF",
        "child_dim_fg": "#7E8BA0",
    },
    "피스타치오": {
        "bg_main": "#F3FAEE",
        "bg_panel": "#FBFEF8",
        "bg_tab": "#E3F2D8",
        "bg_tab_selected": "#CBE8B4",
        "bg_button": "#A8DB7E",
        "bg_button_hover": "#97CC6C",
        "bg_button_pressed": "#81B956",
        "bg_header": "#E7F5DE",
        "text_main": "#243126",
        "text_title": "#456A37",
        "table_select": "#ECF8E4",
        "row_no_bg": "#FAFEF7",
        "child_dim_bg": "#FBFEFA",
        "child_dim_fg": "#7F8D7A",
    },
    "샌드 베이지": {
        "bg_main": "#FBF6EE",
        "bg_panel": "#FEFCF8",
        "bg_tab": "#F3E7D8",
        "bg_tab_selected": "#E8CFB0",
        "bg_button": "#DDBD8F",
        "bg_button_hover": "#CFAE80",
        "bg_button_pressed": "#BA9868",
        "bg_header": "#F4E8D8",
        "text_main": "#2F2B27",
        "text_title": "#7A5A35",
        "table_select": "#F8EEDF",
        "row_no_bg": "#FEFCF8",
        "child_dim_bg": "#FFFDFB",
        "child_dim_fg": "#95887A",
    },
    "슬레이트 그레이": {
        "bg_main": "#F3F5F7",
        "bg_panel": "#FBFCFD",
        "bg_tab": "#E3E8ED",
        "bg_tab_selected": "#CAD3DC",
        "bg_button": "#AAB8C5",
        "bg_button_hover": "#99A9B7",
        "bg_button_pressed": "#8494A2",
        "bg_header": "#E7EBEF",
        "text_main": "#222831",
        "text_title": "#4A5A69",
        "table_select": "#EAEFF3",
        "row_no_bg": "#FAFBFC",
        "child_dim_bg": "#FCFDFD",
        "child_dim_fg": "#7E8A96",
    },
    "코랄 피치": {
        "bg_main": "#FFF4F1",
        "bg_panel": "#FFF9F7",
        "bg_tab": "#FFE3DB",
        "bg_tab_selected": "#FFC4B2",
        "bg_button": "#FF9E80",
        "bg_button_hover": "#F78D6E",
        "bg_button_pressed": "#E17757",
        "bg_header": "#FFE7E0",
        "text_main": "#302624",
        "text_title": "#9A4E3A",
        "table_select": "#FFECE6",
        "row_no_bg": "#FFF9F7",
        "child_dim_bg": "#FFFDFC",
        "child_dim_fg": "#95807A",
    },
}


# =========================
# 공통 유틸
# =========================
def ensure_folder(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_filename(text: str) -> str:
    bad_chars = r'\/:*?"<>|'
    result = text
    for ch in bad_chars:
        result = result.replace(ch, "_")
    return result.strip()


def normalize_text_remove_dash(text: str) -> str:
    if text is None:
        return ""
    return str(text).replace("-", "")


def has_leading_or_trailing_space(text: str) -> bool:
    if text is None:
        return False
    s = str(text)
    return s != s.strip()


def is_row_empty(row_data: dict) -> bool:
    text_exists = any(
        str(row_data.get(k, "")).strip()
        for k in ["문서번호", "요청기관", "실명번호", "계좌번호"]
    )
    check_exists = any(
        bool(row_data.get(k, False))
        for k in ["거래내역", "인적사항", "SMS정보", "거래신청서", "사고신고내역"]
    )
    return not text_exists and not check_exists


def excel_check_to_bool(value) -> bool:
    """
    체크 컬럼 판단 규칙:
    - None, 빈칸, 공백만 있는 값 -> False
    - 그 외 내용이 있으면 True
    """
    if value is None:
        return False

    if isinstance(value, str):
        if value.strip() == "":
            return False
        return True

    return True


# =========================
# 스타일 생성
# =========================
def build_stylesheet(theme):
    return f"""
    QMainWindow, QWidget {{
        background-color: {theme['bg_main']};
        color: {theme['text_main']};
    }}

    QTabWidget::pane {{
        border: 1px solid #D5ECE7;
        background: {theme['bg_panel']};
        border-radius: 16px;
        top: -1px;
    }}

    QTabBar::tab {{
        background: {theme['bg_tab']};
        color: {theme['text_main']};
        border: 1px solid #C9D7D3;
        padding: 10px 22px;
        margin-right: 6px;
        border-top-left-radius: 12px;
        border-top-right-radius: 12px;
        min-width: 120px;
        font-weight: 600;
    }}

    QTabBar::tab:selected {{
        background: {theme['bg_tab_selected']};
        color: {theme['text_main']};
    }}

    QLabel#pageTitle {{
        font-size: 22px;
        font-weight: 700;
        color: {theme['text_title']};
        padding: 2px 0 0 2px;
        background: transparent;
    }}

    QLabel#pageSubtitle {{
        font-size: 11px;
        color: #6B7280;
        padding-left: 2px;
        background: transparent;
    }}

    QFrame#infoCard {{
        background: {theme['bg_panel']};
        border: 1px solid #D4E7E1;
        border-radius: 16px;
    }}

    QLabel#cardTitle {{
        font-size: 14px;
        font-weight: 700;
        color: {theme['text_title']};
        background: transparent;
    }}

    QLabel#cardDesc {{
        font-size: 12px;
        color: #4B5563;
        line-height: 1.5;
        background: transparent;
    }}

    QTableWidget {{
        background: {theme['bg_panel']};
        border: 1px solid #D8E5E0;
        border-radius: 14px;
        gridline-color: #E5E7EB;
        selection-background-color: {theme['table_select']};
        selection-color: {theme['text_main']};
        font-size: 10.5pt;
        outline: none;
    }}

    QHeaderView::section {{
        background-color: {theme['bg_header']};
        color: {theme['text_title']};
        border: none;
        border-bottom: 1px solid #C9DDD7;
        padding: 10px 6px;
        font-weight: 700;
    }}

    QTableWidget::item {{
        padding: 6px;
        border: none;
    }}

    QPushButton {{
        background-color: {theme['bg_button']};
        color: #0E3B36;
        border: none;
        border-radius: 12px;
        padding: 10px 18px;
        min-height: 20px;
        font-size: 10.5pt;
        font-weight: 700;
    }}

    QPushButton:hover {{
        background-color: {theme['bg_button_hover']};
    }}

    QPushButton:pressed {{
        background-color: {theme['bg_button_pressed']};
    }}

    QComboBox {{
        background-color: #FFFFFF;
        color: {theme['text_main']};
        border: 1px solid #C9DDD7;
        border-radius: 10px;
        padding: 6px 10px;
        min-width: 140px;
    }}

    QComboBox::drop-down {{
        border: none;
        width: 24px;
    }}

    QMessageBox {{
        background-color: #FFFFFF;
    }}
    """


# =========================
# 테이블 정렬 delegate
# =========================
class CenterAlignDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = Qt.AlignCenter


# =========================
# 카드 UI
# =========================
class InfoCard(QFrame):
    def __init__(self, title: str, description: str, parent=None):
        super().__init__(parent)
        self.setObjectName("infoCard")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("cardTitle")

        desc_label = QLabel(description)
        desc_label.setObjectName("cardDesc")
        desc_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(desc_label)


# =========================
# 작성/취합 공통 테이블
# =========================
class BaseTable(QTableWidget):
    def __init__(self, row_count=10, parent=None):
        super().__init__(row_count, len(COLUMNS), parent)
        self.center_delegate = CenterAlignDelegate()
        self.setItemDelegate(self.center_delegate)
        self.theme = THEMES["민트 아쿠아"]
        self.setup_ui()
        self.fill_row_numbers()

    def set_theme(self, theme):
        self.theme = theme
        self.refresh_base_colors()

    def setup_ui(self):
        self.setHorizontalHeaderLabels(COLUMNS)
        self.verticalHeader().setVisible(False)
        self.setAlternatingRowColors(False)
        self.setShowGrid(False)
        self.setSelectionBehavior(QTableWidget.SelectItems)
        self.setSelectionMode(QTableWidget.SingleSelection)
        self.setFocusPolicy(Qt.StrongFocus)

        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Interactive)
        header.setSectionResizeMode(2, QHeaderView.Interactive)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        header.setSectionResizeMode(4, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        header.setSectionResizeMode(6, QHeaderView.Fixed)
        header.setSectionResizeMode(7, QHeaderView.Fixed)
        header.setSectionResizeMode(8, QHeaderView.Fixed)
        header.setSectionResizeMode(9, QHeaderView.Fixed)
        header.setSectionResizeMode(10, QHeaderView.Fixed)

        self.setColumnWidth(0, 60)
        self.setColumnWidth(1, 160)
        self.setColumnWidth(2, 180)
        self.setColumnWidth(3, 170)
        self.setColumnWidth(4, 170)
        self.setColumnWidth(5, 95)
        self.setColumnWidth(6, 95)
        self.setColumnWidth(7, 95)
        self.setColumnWidth(8, 110)
        self.setColumnWidth(9, 120)
        self.setColumnWidth(10, 110)

        self.setMinimumHeight(520)

        for row in range(self.rowCount()):
            self._create_empty_row(row)

    def _create_empty_row(self, row: int):
        no_item = QTableWidgetItem(str(row + 1))
        no_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        no_item.setTextAlignment(Qt.AlignCenter)
        self.setItem(row, 0, no_item)

        for col in TEXT_COLUMNS:
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            self.setItem(row, col, item)

        for col in CHECK_COLUMNS:
            item = QTableWidgetItem(CHECK_OFF)
            item.setTextAlignment(Qt.AlignCenter)
            self.setItem(row, col, item)

        self.apply_base_row_colors(row)

    def apply_base_row_colors(self, row: int):
        for col in range(self.columnCount()):
            item = self.item(row, col)
            if item is None:
                continue

            if col == 0:
                item.setBackground(QBrush(QColor(self.theme["row_no_bg"])))
                item.setForeground(QBrush(QColor(self.theme["text_main"])))
            else:
                item.setBackground(QBrush(QColor("#FFFFFF")))
                item.setForeground(QBrush(QColor(self.theme["text_main"])))

    def refresh_base_colors(self):
        for row in range(self.rowCount()):
            self.apply_base_row_colors(row)

    def fill_row_numbers(self):
        for row in range(self.rowCount()):
            item = self.item(row, 0)
            if item is None:
                item = QTableWidgetItem()
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.setItem(row, 0, item)
            item.setText(str(row + 1))
            item.setTextAlignment(Qt.AlignCenter)

    def add_empty_row(self):
        row = self.rowCount()
        self.insertRow(row)
        self._create_empty_row(row)
        self.fill_row_numbers()

    def clear_all_rows(self, row_count=10):
        self.setRowCount(0)
        self.setRowCount(row_count)
        for row in range(row_count):
            self._create_empty_row(row)
        self.fill_row_numbers()

    def set_check_value(self, row: int, col: int, checked: bool):
        item = self.item(row, col)
        if item is None:
            item = QTableWidgetItem()
            self.setItem(row, col, item)
        item.setText(CHECK_ON if checked else CHECK_OFF)
        item.setTextAlignment(Qt.AlignCenter)

    def get_check_value(self, row: int, col: int) -> bool:
        item = self.item(row, col)
        if item is None:
            return False
        return item.text() == CHECK_ON

    def toggle_check_item(self, row: int, col: int):
        if col not in CHECK_COLUMNS:
            return
        current = self.get_check_value(row, col)
        self.set_check_value(row, col, not current)

    def adjust_doc_org_columns(self):
        fm = self.fontMetrics()

        max_doc_width = fm.horizontalAdvance("문서번호") + 36
        max_org_width = fm.horizontalAdvance("요청기관") + 36

        for row in range(self.rowCount()):
            item_doc = self.item(row, 1)
            item_org = self.item(row, 2)

            if item_doc:
                text = item_doc.text() or ""
                width = fm.horizontalAdvance(text) + 36
                if width > max_doc_width:
                    max_doc_width = width

            if item_org:
                text = item_org.text() or ""
                width = fm.horizontalAdvance(text) + 36
                if width > max_org_width:
                    max_org_width = width

        max_doc_width = max(140, min(max_doc_width, 420))
        max_org_width = max(160, min(max_org_width, 520))

        self.setColumnWidth(1, max_doc_width)
        self.setColumnWidth(2, max_org_width)


# =========================
# 작성탭 테이블
# =========================
class WriteTable(BaseTable):
    def __init__(self, row_count=10, parent=None):
        self.group_row_map = {}
        self.loading_mode = False
        super().__init__(row_count, parent)
        self.itemChanged.connect(self.on_item_changed)
        self.apply_all_row_styles()

    def set_theme(self, theme):
        super().set_theme(theme)
        self.apply_all_row_styles()

    def apply_all_row_styles(self):
        for row in range(self.rowCount()):
            self.apply_row_style(row)

    def apply_row_style(self, row: int):
        is_child = self.is_child_row(row)

        for col in range(self.columnCount()):
            item = self.item(row, col)
            if item is None:
                continue

            if col == 0:
                item.setBackground(QBrush(QColor(self.theme["row_no_bg"])))
                item.setForeground(QBrush(QColor(self.theme["text_main"])))
                continue

            if is_child:
                if col in [1, 2]:
                    item.setBackground(QBrush(QColor(self.theme["child_dim_bg"])))
                    item.setForeground(QBrush(QColor(self.theme["child_dim_fg"])))
                else:
                    item.setBackground(QBrush(QColor("#FFFFFF")))
                    item.setForeground(QBrush(QColor(self.theme["text_main"])))
            else:
                item.setBackground(QBrush(QColor("#FFFFFF")))
                item.setForeground(QBrush(QColor(self.theme["text_main"])))

    def get_row_group_key(self, row: int) -> str:
        doc_no = self.item(row, 1).text().strip() if self.item(row, 1) else ""
        org = self.item(row, 2).text().strip() if self.item(row, 2) else ""
        if not doc_no or not org:
            return ""
        return f"{doc_no}||{org}"

    def is_child_row(self, row: int) -> bool:
        group_key = self.get_row_group_key(row)
        if not group_key:
            return False
        parent_row = self.group_row_map.get(group_key)
        return parent_row is not None and parent_row != row

    def on_item_changed(self, item):
        if self.loading_mode:
            return

        row = item.row()
        col = item.column()

        if col in [1, 2]:
            self.refresh_group_map()
            self.sync_children_under_parent(row)
            self.apply_all_row_styles()
            self.adjust_doc_org_columns()

    def refresh_group_map(self):
        self.group_row_map = {}
        for row in range(self.rowCount()):
            doc_no = self.item(row, 1).text().strip() if self.item(row, 1) else ""
            org = self.item(row, 2).text().strip() if self.item(row, 2) else ""
            if doc_no and org:
                key = f"{doc_no}||{org}"
                if key not in self.group_row_map:
                    self.group_row_map[key] = row

    def sync_children_under_parent(self, parent_row: int):
        doc_no = self.item(parent_row, 1).text().strip() if self.item(parent_row, 1) else ""
        org = self.item(parent_row, 2).text().strip() if self.item(parent_row, 2) else ""
        if not doc_no or not org:
            return

        key = f"{doc_no}||{org}"
        if self.group_row_map.get(key) != parent_row:
            return

        for row in range(parent_row + 1, self.rowCount()):
            row_doc = self.item(row, 1).text().strip() if self.item(row, 1) else ""
            row_org = self.item(row, 2).text().strip() if self.item(row, 2) else ""

            if row_doc == doc_no and row_org == org:
                self.loading_mode = True
                self.item(row, 1).setText(doc_no)
                self.item(row, 2).setText(org)
                self.loading_mode = False
                self.apply_row_style(row)

    def get_row_data(self, row: int) -> dict:
        return {
            "순번": row + 1,
            "문서번호": self.item(row, 1).text() if self.item(row, 1) else "",
            "요청기관": self.item(row, 2).text() if self.item(row, 2) else "",
            "실명번호": self.item(row, 3).text() if self.item(row, 3) else "",
            "계좌번호": self.item(row, 4).text() if self.item(row, 4) else "",
            "거래내역": self.get_check_value(row, 5),
            "인적사항": self.get_check_value(row, 6),
            "SMS정보": self.get_check_value(row, 7),
            "거래신청서": self.get_check_value(row, 8),
            "사고신고내역": self.get_check_value(row, 9),
            "담당자명": self.item(row, 10).text() if self.item(row, 10) else "",
        }

    def get_all_non_empty_rows(self) -> list:
        rows = []
        for row in range(self.rowCount()):
            row_data = self.get_row_data(row)
            if not is_row_empty(row_data):
                rows.append(row_data)
        return rows

    def set_row_data(self, row: int, data: dict):
        while row >= self.rowCount():
            self.add_empty_row()

        mapping = {
            1: "문서번호",
            2: "요청기관",
            3: "실명번호",
            4: "계좌번호",
            10: "담당자명",
        }

        for col, key in mapping.items():
            item = self.item(row, col)
            if item is None:
                item = QTableWidgetItem("")
                self.setItem(row, col, item)
            item.setText(str(data.get(key, "")))
            item.setTextAlignment(Qt.AlignCenter)

        check_mapping = {
            5: "거래내역",
            6: "인적사항",
            7: "SMS정보",
            8: "거래신청서",
            9: "사고신고내역",
        }

        for col, key in check_mapping.items():
            self.set_check_value(row, col, bool(data.get(key, False)))

        self.apply_row_style(row)

    def load_rows(self, rows: list):
        self.loading_mode = True
        self.clear_all_rows(max(len(rows), DEFAULT_WRITE_ROWS))
        for idx, row_data in enumerate(rows):
            self.set_row_data(idx, row_data)
        self.loading_mode = False
        self.refresh_group_map()
        self.apply_all_row_styles()
        self.adjust_doc_org_columns()

    def remove_empty_rows(self):
        for row in reversed(range(self.rowCount())):
            row_data = self.get_row_data(row)
            if is_row_empty(row_data):
                self.removeRow(row)

        if self.rowCount() == 0:
            self.add_empty_row()

        self.fill_row_numbers()
        self.refresh_group_map()
        self.apply_all_row_styles()
        self.adjust_doc_org_columns()

    def mousePressEvent(self, event):
        item = self.itemAt(event.pos())
        if item:
            row = item.row()
            col = item.column()
            if col in CHECK_COLUMNS:
                self.toggle_check_item(row, col)
                return
        super().mousePressEvent(event)

    def keyPressEvent(self, event):
        key = event.key()
        current_row = self.currentRow()
        current_col = self.currentColumn()

        if key == Qt.Key_Space and current_col in CHECK_COLUMNS and current_row >= 0:
            self.toggle_check_item(current_row, current_col)
            event.accept()
            return

        if key in (Qt.Key_Return, Qt.Key_Enter):
            next_row = current_row + 1
            if next_row >= self.rowCount():
                self.add_empty_row()
            self.setCurrentCell(next_row, current_col if current_col >= 1 else 1)
            event.accept()
            return

        super().keyPressEvent(event)


# =========================
# 취합탭 테이블
# =========================
class CollectTable(BaseTable):
    def __init__(self, row_count=10, parent=None):
        super().__init__(row_count, parent)
        self.setEditTriggers(QTableWidget.NoEditTriggers)

    def clear_and_load(self, rows: list):
        self.clear_all_rows(max(len(rows), 10))

        for row_idx, row_data in enumerate(rows):
            values = [
                str(row_idx + 1),
                str(row_data.get("문서번호", "")),
                str(row_data.get("요청기관", "")),
                str(row_data.get("실명번호", "")),
                str(row_data.get("계좌번호", "")),
            ]

            for col_idx, value in enumerate(values):
                item = self.item(row_idx, col_idx)
                if item is None:
                    item = QTableWidgetItem("")
                    self.setItem(row_idx, col_idx, item)
                item.setText(value)
                item.setTextAlignment(Qt.AlignCenter)
                if col_idx == 0:
                    item.setBackground(QBrush(QColor(self.theme["row_no_bg"])))
                else:
                    item.setBackground(QBrush(QColor("#FFFFFF")))
                item.setForeground(QBrush(QColor(self.theme["text_main"])))

            self.set_check_value(row_idx, 5, row_data.get("거래내역", False))
            self.set_check_value(row_idx, 6, row_data.get("인적사항", False))
            self.set_check_value(row_idx, 7, row_data.get("SMS정보", False))
            self.set_check_value(row_idx, 8, row_data.get("거래신청서", False))
            self.set_check_value(row_idx, 9, row_data.get("사고신고내역", False))

            for chk_col in CHECK_COLUMNS:
                chk_item = self.item(row_idx, chk_col)
                if chk_item:
                    chk_item.setBackground(QBrush(QColor("#FFFFFF")))
                    chk_item.setForeground(QBrush(QColor(self.theme["text_main"])))

            item_writer = self.item(row_idx, 10)
            if item_writer is None:
                item_writer = QTableWidgetItem("")
                self.setItem(row_idx, 10, item_writer)
            item_writer.setText(str(row_data.get("담당자명", row_data.get("_writer", ""))))
            item_writer.setTextAlignment(Qt.AlignCenter)
            item_writer.setBackground(QBrush(QColor("#FFFFFF")))
            item_writer.setForeground(QBrush(QColor(self.theme["text_main"])))

        self.adjust_doc_org_columns()


# =========================
# 작성 탭
# =========================
class WriteTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.writer_name = getpass.getuser()
        self.current_theme = THEMES["민트 아쿠아"]
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(22, 22, 22, 22)
        main_layout.setSpacing(16)

        title = QLabel("금융거래정보 수집 전산화 기초자료 작성")
        title.setObjectName("pageTitle")

        subtitle = QLabel(f"작성자 : {self.writer_name}    |    저장 위치 : {SHARED_FOLDER}")
        subtitle.setObjectName("pageSubtitle")

        info = InfoCard(
            "작성탭",
            "문서번호와 요청기관은 표 안에서 입력합니다. '+ 명의인/계좌 추가'는 현재 선택한 문서번호 묶음 아래에 "
            "하위 항목처럼 추가되며, 저장 시 같은 문서번호와 요청기관이 모든 관련 행에 함께 반영됩니다. "
            "엑셀 불러오기는 문서번호부터 사고신고내역까지 모두 읽어오며, 체크 컬럼은 내용이 있으면 체크로 처리합니다."
        )

        self.table = WriteTable(row_count=DEFAULT_WRITE_ROWS)
        self.table.set_theme(self.current_theme)
        for row in range(self.table.rowCount()):
            self.table.item(row, 10).setText(self.writer_name)
        self.table.adjust_doc_org_columns()

        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        self.btn_add_row = QPushButton("행 추가")
        self.btn_add_child = QPushButton("+ 명의인/계좌 추가")
        self.btn_remove_empty = QPushButton("빈 행 삭제")
        self.btn_import_excel = QPushButton("엑셀 불러오기")
        self.btn_load = QPushButton("문서번호 불러오기")
        self.btn_save = QPushButton("저장")

        self.btn_add_row.clicked.connect(self.add_row)
        self.btn_add_child.clicked.connect(self.add_child_row)
        self.btn_remove_empty.clicked.connect(self.remove_empty_rows)
        self.btn_import_excel.clicked.connect(self.import_excel)
        self.btn_load.clicked.connect(self.load_by_doc_no)
        self.btn_save.clicked.connect(self.save_data)

        button_layout.addWidget(self.btn_add_row)
        button_layout.addWidget(self.btn_add_child)
        button_layout.addWidget(self.btn_remove_empty)
        button_layout.addWidget(self.btn_import_excel)
        button_layout.addWidget(self.btn_load)
        button_layout.addStretch()
        button_layout.addWidget(self.btn_save)

        main_layout.addWidget(title)
        main_layout.addWidget(subtitle)
        main_layout.addWidget(info)
        main_layout.addWidget(self.table)
        main_layout.addLayout(button_layout)

    def set_theme(self, theme):
        self.current_theme = theme
        self.table.set_theme(theme)
        self.table.apply_all_row_styles()

    def add_row(self):
        self.table.add_empty_row()
        self.table.item(self.table.rowCount() - 1, 10).setText(self.writer_name)
        self.table.apply_all_row_styles()
        self.table.adjust_doc_org_columns()

    def add_child_row(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "추가 불가", "먼저 기준이 될 행을 선택해 주세요.")
            return

        doc_no = self.table.item(current_row, 1).text().strip() if self.table.item(current_row, 1) else ""
        org = self.table.item(current_row, 2).text().strip() if self.table.item(current_row, 2) else ""

        if not doc_no or not org:
            QMessageBox.warning(self, "추가 불가", "문서번호와 요청기관이 입력된 행에서만 명의인/계좌를 추가할 수 있습니다.")
            return

        insert_row = current_row + 1
        for row in range(current_row + 1, self.table.rowCount()):
            row_doc = self.table.item(row, 1).text().strip() if self.table.item(row, 1) else ""
            row_org = self.table.item(row, 2).text().strip() if self.table.item(row, 2) else ""
            if row_doc == doc_no and row_org == org:
                insert_row = row + 1
            else:
                break

        self.table.insertRow(insert_row)
        self.table._create_empty_row(insert_row)

        self.table.loading_mode = True
        self.table.item(insert_row, 1).setText(doc_no)
        self.table.item(insert_row, 2).setText(org)
        self.table.item(insert_row, 10).setText(self.writer_name)
        self.table.loading_mode = False

        self.table.fill_row_numbers()
        self.table.refresh_group_map()
        self.table.apply_all_row_styles()
        self.table.adjust_doc_org_columns()
        self.table.setCurrentCell(insert_row, 3)

    def remove_empty_rows(self):
        self.table.remove_empty_rows()

    def import_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "엑셀 파일 선택",
            "",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            imported_rows = []
            # 1행 = 헤더, 2행부터 데이터
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                if excel_row is None:
                    continue

                doc_no = excel_row[0] if len(excel_row) > 0 and excel_row[0] is not None else ""
                org = excel_row[1] if len(excel_row) > 1 and excel_row[1] is not None else ""
                real_id = excel_row[2] if len(excel_row) > 2 and excel_row[2] is not None else ""
                acc_no = excel_row[3] if len(excel_row) > 3 and excel_row[3] is not None else ""

                chk1_raw = excel_row[4] if len(excel_row) > 4 else None
                chk2_raw = excel_row[5] if len(excel_row) > 5 else None
                chk3_raw = excel_row[6] if len(excel_row) > 6 else None
                chk4_raw = excel_row[7] if len(excel_row) > 7 else None
                chk5_raw = excel_row[8] if len(excel_row) > 8 else None

                writer_name = excel_row[9] if len(excel_row) > 9 and excel_row[9] is not None else self.writer_name

                row_data = {
                    "문서번호": str(doc_no),
                    "요청기관": str(org),
                    "실명번호": str(real_id),
                    "계좌번호": str(acc_no),
                    "거래내역": excel_check_to_bool(chk1_raw),
                    "인적사항": excel_check_to_bool(chk2_raw),
                    "SMS정보": excel_check_to_bool(chk3_raw),
                    "거래신청서": excel_check_to_bool(chk4_raw),
                    "사고신고내역": excel_check_to_bool(chk5_raw),
                    "담당자명": str(writer_name),
                }

                if not is_row_empty(row_data):
                    imported_rows.append(row_data)

            if not imported_rows:
                QMessageBox.information(self, "불러오기 결과", "불러올 데이터가 없습니다.")
                return

            self.table.load_rows(imported_rows)

            # 담당자명이 비어 있으면 현재 사용자명 보정
            for row in range(self.table.rowCount()):
                item = self.table.item(row, 10)
                if item and item.text().strip() == "":
                    item.setText(self.writer_name)

            self.table.refresh_group_map()
            self.table.apply_all_row_styles()
            self.table.adjust_doc_org_columns()

            QMessageBox.information(self, "불러오기 완료", f"엑셀에서 {len(imported_rows)}건을 불러왔습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 불러오기 중 오류가 발생했습니다.\n\n{str(e)}")

    def validate_and_normalize_before_save(self):
        rows = self.table.get_all_non_empty_rows()
        if not rows:
            QMessageBox.warning(self, "저장 불가", "저장할 데이터가 없습니다.")
            return None

        normalized_rows = []

        for row in rows:
            doc_no_raw = row.get("문서번호", "")
            org_raw = row.get("요청기관", "")
            id_raw = row.get("실명번호", "")
            acc_raw = row.get("계좌번호", "")

            fields_to_check = [doc_no_raw, org_raw, id_raw, acc_raw]
            if any(has_leading_or_trailing_space(x) for x in fields_to_check):
                QMessageBox.warning(
                    self,
                    "입력 오류",
                    "입력하신 데이터 중 공백으로 시작 또는 끝나는 행이 있습니다. 수정 후 다시 저장하시기 바랍니다"
                )
                return None

            normalized_row = {
                "문서번호": normalize_text_remove_dash(str(doc_no_raw).strip()),
                "요청기관": normalize_text_remove_dash(str(org_raw).strip()),
                "실명번호": normalize_text_remove_dash(str(id_raw).strip()),
                "계좌번호": normalize_text_remove_dash(str(acc_raw).strip()),
                "거래내역": bool(row.get("거래내역", False)),
                "인적사항": bool(row.get("인적사항", False)),
                "SMS정보": bool(row.get("SMS정보", False)),
                "거래신청서": bool(row.get("거래신청서", False)),
                "사고신고내역": bool(row.get("사고신고내역", False)),
                "담당자명": str(row.get("담당자명", "")).strip() or self.writer_name,
            }

            if not normalized_row["문서번호"] or not normalized_row["요청기관"]:
                QMessageBox.warning(self, "입력 오류", "문서번호와 요청기관은 반드시 입력해야 합니다.")
                return None

            normalized_rows.append(normalized_row)

        final_rows = []
        current_doc = ""
        current_org = ""

        for row in normalized_rows:
            if row["문서번호"] and row["요청기관"]:
                current_doc = row["문서번호"]
                current_org = row["요청기관"]
            else:
                row["문서번호"] = current_doc
                row["요청기관"] = current_org

            final_rows.append(row)

        self.table.adjust_doc_org_columns()
        return final_rows

    def save_data(self):
        ensure_folder(SHARED_FOLDER)

        rows = self.validate_and_normalize_before_save()
        if rows is None:
            return

        doc_group_map = {}
        for row in rows:
            key = row["문서번호"]
            doc_group_map.setdefault(key, []).append(row)

        try:
            for doc_no, doc_rows in doc_group_map.items():
                org = doc_rows[0].get("요청기관", "")
                writer_name = doc_rows[0].get("담당자명", self.writer_name)

                payload = {
                    "writer": writer_name,
                    "saved_at": now_str(),
                    "문서번호": doc_no,
                    "요청기관": org,
                    "rows": doc_rows
                }

                save_path = os.path.join(SHARED_FOLDER, f"{safe_filename(doc_no)}.json")
                with open(save_path, "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)

            QMessageBox.information(
                self,
                "저장 완료",
                "저장되었습니다.\n같은 문서번호 파일이 있으면 덮어쓰기 처리되었습니다."
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 중 오류가 발생했습니다.\n\n{str(e)}")

    def load_by_doc_no(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "불러오기 불가", "먼저 문서번호가 입력된 행을 선택해 주세요.")
            return

        doc_no_raw = self.table.item(current_row, 1).text() if self.table.item(current_row, 1) else ""
        if has_leading_or_trailing_space(doc_no_raw):
            QMessageBox.warning(
                self,
                "입력 오류",
                "입력하신 데이터 중 공백으로 시작 또는 끝나는 행이 있습니다. 수정 후 다시 저장하시기 바랍니다"
            )
            return

        doc_no = normalize_text_remove_dash(str(doc_no_raw).strip())
        if not doc_no:
            QMessageBox.warning(self, "불러오기 불가", "선택한 행의 문서번호를 먼저 입력해 주세요.")
            return

        load_path = os.path.join(SHARED_FOLDER, f"{safe_filename(doc_no)}.json")
        if not os.path.exists(load_path):
            QMessageBox.information(self, "조회 결과", "해당 문서번호로 저장된 파일이 없습니다.")
            return

        try:
            with open(load_path, "r", encoding="utf-8") as f:
                payload = json.load(f)

            rows = payload.get("rows", [])
            self.table.load_rows(rows)
            self.table.adjust_doc_org_columns()
            QMessageBox.information(self, "불러오기 완료", f"{doc_no} 문서번호 데이터를 불러왔습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"불러오기 중 오류가 발생했습니다.\n\n{str(e)}")


# =========================
# 취합 탭
# =========================
class CollectTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.loaded_rows = []
        self.current_theme = THEMES["민트 아쿠아"]
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(22, 22, 22, 22)
        main_layout.setSpacing(16)

        title = QLabel("금융거래정보 수집 전산화 기초자료 취합")
        title.setObjectName("pageTitle")

        subtitle = QLabel(f"불러오기 위치 : {SHARED_FOLDER}")
        subtitle.setObjectName("pageSubtitle")

        info = InfoCard(
            "취합탭",
            "공유폴더의 JSON 파일을 모두 불러옵니다. 같은 문서번호에 여러 명의인/계좌가 있으면 "
            "엑셀 저장 시 각 행에 동일한 문서번호와 요청기관이 반복 입력됩니다."
        )

        self.table = CollectTable(row_count=10)
        self.table.set_theme(self.current_theme)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        self.btn_load = QPushButton("불러오기")
        self.btn_refresh = QPushButton("새로고침")
        self.btn_export = QPushButton("취합 저장(.xlsx)")

        self.btn_load.clicked.connect(self.load_data)
        self.btn_refresh.clicked.connect(self.load_data)
        self.btn_export.clicked.connect(self.export_to_excel)

        button_layout.addWidget(self.btn_load)
        button_layout.addWidget(self.btn_refresh)
        button_layout.addStretch()
        button_layout.addWidget(self.btn_export)

        main_layout.addWidget(title)
        main_layout.addWidget(subtitle)
        main_layout.addWidget(info)
        main_layout.addWidget(self.table)
        main_layout.addLayout(button_layout)

    def set_theme(self, theme):
        self.current_theme = theme
        self.table.set_theme(theme)
        if self.loaded_rows:
            self.table.clear_and_load(self.loaded_rows)
        else:
            self.table.refresh_base_colors()

    def load_data(self):
        ensure_folder(SHARED_FOLDER)

        files = [f for f in os.listdir(SHARED_FOLDER) if f.lower().endswith(".json")]
        all_rows = []

        for file_name in sorted(files):
            full_path = os.path.join(SHARED_FOLDER, file_name)
            try:
                with open(full_path, "r", encoding="utf-8") as f:
                    payload = json.load(f)

                doc_no = payload.get("문서번호", "")
                org = payload.get("요청기관", "")
                writer = payload.get("writer", "")
                saved_at = payload.get("saved_at", "")
                rows = payload.get("rows", [])

                for row in rows:
                    row_copy = {
                        "문서번호": row.get("문서번호", doc_no),
                        "요청기관": row.get("요청기관", org),
                        "실명번호": row.get("실명번호", ""),
                        "계좌번호": row.get("계좌번호", ""),
                        "거래내역": row.get("거래내역", False),
                        "인적사항": row.get("인적사항", False),
                        "SMS정보": row.get("SMS정보", False),
                        "거래신청서": row.get("거래신청서", False),
                        "사고신고내역": row.get("사고신고내역", False),
                        "담당자명": row.get("담당자명", writer),
                        "_writer": writer,
                        "_saved_at": saved_at,
                    }
                    all_rows.append(row_copy)
            except Exception:
                continue

        self.loaded_rows = all_rows
        self.table.clear_and_load(all_rows)

        if not all_rows:
            QMessageBox.information(self, "불러오기 결과", "불러올 데이터가 없습니다.")
        else:
            QMessageBox.information(self, "불러오기 완료", f"{len(all_rows)}건을 불러왔습니다.")

    def export_to_excel(self):
        if not self.loaded_rows:
            QMessageBox.warning(self, "저장 불가", "먼저 데이터를 불러와 주세요.")
            return

        default_name = f"금융거래정보_기초자료_취합_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "취합 엑셀 저장",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "취합자료"

            excel_headers = [
                "순번", "문서번호", "요청기관", "실명번호", "계좌번호",
                "거래내역", "인적사항", "SMS정보", "거래신청서", "사고신고내역",
                "담당자명", "저장일시"
            ]
            ws.append(excel_headers)

            header_fill = PatternFill(fill_type="solid", fgColor="B7F0E8")
            header_font = XLFont(bold=True, color="1F2937")
            thin = Side(style="thin", color="D1D5DB")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for idx, row in enumerate(self.loaded_rows, start=1):
                excel_row = [
                    idx,
                    row.get("문서번호", ""),
                    row.get("요청기관", ""),
                    row.get("실명번호", ""),
                    row.get("계좌번호", ""),
                    "Y" if row.get("거래내역", False) else "",
                    "Y" if row.get("인적사항", False) else "",
                    "Y" if row.get("SMS정보", False) else "",
                    "Y" if row.get("거래신청서", False) else "",
                    "Y" if row.get("사고신고내역", False) else "",
                    row.get("담당자명", row.get("_writer", "")),
                    row.get("_saved_at", ""),
                ]
                ws.append(excel_row)

            widths = {
                "A": 8,
                "B": 22,
                "C": 30,
                "D": 20,
                "E": 20,
                "F": 12,
                "G": 12,
                "H": 12,
                "I": 14,
                "J": 16,
                "K": 14,
                "L": 20
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            wb.save(save_path)

            QMessageBox.information(self, "저장 완료", f".xlsx 파일로 저장되었습니다.\n\n{save_path}")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류가 발생했습니다.\n\n{str(e)}")


# =========================
# 메인 윈도우
# =========================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        ensure_folder(SHARED_FOLDER)

        self.current_theme_name = "민트 아쿠아"
        self.current_theme = THEMES[self.current_theme_name]

        self.setWindowTitle("금융거래정보 수집 전산화 기초자료 작성/취합")
        self.resize(1640, 900)
        self.setup_ui()
        self.apply_theme(self.current_theme_name)

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        layout = QVBoxLayout(central)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        top_bar = QHBoxLayout()
        top_bar.setSpacing(10)

        theme_label = QLabel("테마 선택")
        theme_label.setObjectName("pageSubtitle")

        self.theme_combo = QComboBox()
        self.theme_combo.addItems(list(THEMES.keys()))
        self.theme_combo.setCurrentText(self.current_theme_name)
        self.theme_combo.currentTextChanged.connect(self.apply_theme)

        top_bar.addStretch()
        top_bar.addWidget(theme_label)
        top_bar.addWidget(self.theme_combo)

        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        self.tabs.setMovable(False)

        self.write_tab = WriteTab()
        self.collect_tab = CollectTab()

        self.tabs.addTab(self.write_tab, "작성탭")
        self.tabs.addTab(self.collect_tab, "취합탭")

        layout.addLayout(top_bar)
        layout.addWidget(self.tabs)

    def apply_theme(self, theme_name):
        if theme_name not in THEMES:
            return

        self.current_theme_name = theme_name
        self.current_theme = THEMES[theme_name]

        QApplication.instance().setFont(QFont("맑은 고딕", 10))
        self.setStyleSheet(build_stylesheet(self.current_theme))

        self.write_tab.set_theme(self.current_theme)
        self.collect_tab.set_theme(self.current_theme)


# =========================
# 실행
# =========================
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()